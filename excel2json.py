from collections import namedtuple
from datetime import datetime

import openpyxl

SUB_OBJECT = '__'


class WookbookTemplate(namedtuple('WookbookTemplate', 'sheets')):
    def excel(self):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.active)
        for sheet in self.sheets:
            sheet.excel(wb)
        return wb

    def write(self, filename):
        wb = self.excel()
        wb.save(filename=filename)

    def json(self, excel_workbook):
        json = {}
        for sheet in self.sheets:
            json[sheet.name] = sheet.json(excel_workbook)
        return json


    def read(self, filename):
        return openpyxl.load_workbook(filename)

class Sheet(namedtuple('Sheet', 'name columns')):
    def excel(self, excel_workbook):
        ws = excel_workbook.create_sheet(title=self.name)
        for n, column in enumerate(self.columns):
            column.excel(ws, n + 1)

    def json(self, excel_workbook):
        ws = excel_workbook.get_sheet_by_name(self.name)
        json = []
        sub_columns = [c for c in self.columns if c.sub_object_name is not None]
        obj = None
        for r in range(1, len(ws.rows)):
            values = tuple([col.json(ws, c, r) for c, col in enumerate(self.columns)])
            items = zip(self.columns, values)
            obj_values = tuple([v for c, v in items if c.sub_object_name is None])
            if obj_values != tuple([None] * len(obj_values)):
                obj = dict([(c.name, v) for c, v in items if c.sub_object_name is None])
                json.append(obj)
            if obj is not None:
                sub_keys = set([c.sub_object_name for c in self.columns if c.sub_object_name is not None])
                for sub_key in sub_keys:
                    if sub_key not in obj:
                        obj[sub_key] = []
                    sub_item = dict([(c.sub_object_key, v) for c,v in items if c.name.startswith(sub_key)])
                    obj[sub_key].append(dict(sub_item))
        return json


class Column(namedtuple('Column', 'name type')):
    def excel(self, excel_worksheet, index):
        excel_worksheet.cell(column=index, row=1).value = self.name

    def json(self, excel_worksheet, index, row):
        return excel_worksheet.cell(column=index + 1, row=row + 1).value

    @property
    def sub_object_name(self):
        return self.name.partition(SUB_OBJECT)[0] if SUB_OBJECT in self.name else None
    
    @property
    def sub_object_key(self):
        return self.name.partition(SUB_OBJECT)[2] if SUB_OBJECT in self.name else None


WORKBOOK_TEMPLATES = dict(
    site=WookbookTemplate((
        Sheet(
            name='nav',
            columns=(
                Column('name', unicode),
                Column('target', unicode),
            )
        ),
        Sheet(
            name='footer_columns',
            columns=(
                Column('section', unicode),
                Column('name', unicode),
                Column('target', unicode),
            ),
        ),
        Sheet(
            name='footer_links',
            columns=(
                Column('section', unicode),
                Column('top_link', unicode),
                Column('links%sname' % SUB_OBJECT, unicode),
                Column('links%starget' % SUB_OBJECT, unicode),
                Column('links%simage' % SUB_OBJECT, unicode),
            ),
        ),
        Sheet(
            name='jumbotron',
            columns=(
                Column('title', unicode),
                Column('expires', datetime),
                Column('background_image', unicode),
                Column('description', unicode),
                Column('actions%stext' % SUB_OBJECT, unicode),
                Column('actions%starget' % SUB_OBJECT, unicode),
            ),
        ),
    )),
    events=WookbookTemplate((
        Sheet(
            name='events',
            columns=(
                Column('name', unicode),
                Column('start', datetime),
                Column('end', datetime),
                Column('description', unicode),
                Column('image', unicode),
                Column('actions%stext' % SUB_OBJECT, unicode),
                Column('actions%starget' % SUB_OBJECT, unicode)
            ),
        ),
    )),
    pages=WookbookTemplate((
        Sheet(
            name='pages',
            columns=(
                Column('url', unicode),
                Column('title', unicode),
                Column('text', unicode),
                Column('image', unicode),
            ),
        ),
    )),
)



if __name__ == "__main__":
    for name, workbook in WORKBOOK_TEMPLATES.iteritems():
        filename = "static/%s.xlsx" % name
        workbook.write(filename)
        print "Wrote %s" % filename
